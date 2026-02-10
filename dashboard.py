import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import os

# ページ設定
st.set_page_config(
    page_title="工数管理ダッシュボード",
    page_icon="📊",
    layout="wide"
)

# 日本語フォント設定
plt.rcParams['font.sans-serif'] = ['Yu Gothic', 'MS Gothic', 'Hiragino Sans', 'IPAexGothic']
plt.rcParams['axes.unicode_minus'] = False

# Excelファイルパス（固定）
LOG_FILE = r"C:\Users\akasaka.kazuyuki\OneDrive - ユーザーサイド株式会社\strat-lab\strat-lab\10_Daily\11_工数管理\Pythonログ\work_log.xlsx"

# カスタムCSS
st.markdown("""
<style>
    .stApp {
        background-color: #1a1a2e;
    }
    h1, h2, h3 {
        color: #f1f5f9 !important;
    }
    .stDateInput label, .stRadio label {
        color: #94a3b8 !important;
    }
</style>
""", unsafe_allow_html=True)

# タイトル
st.title("📊 工数管理ダッシュボード")
st.markdown("---")

# データ読み込み
@st.cache_data(ttl=60)  # 60秒キャッシュ
def load_data():
    if not os.path.exists(LOG_FILE):
        return pd.DataFrame()
    
    wb = load_workbook(LOG_FILE)
    ws = wb.active
    data = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            try:
                date_str = str(row[0]).split()[0]
                for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
                    try:
                        row_date = datetime.strptime(date_str, fmt).date()
                        break
                    except:
                        continue
                else:
                    continue
                
                data.append({
                    "日付": row_date,
                    "開始": row[1],
                    "終了": row[2],
                    "タスク": row[3],
                    "分": float(row[4]) if row[4] else 0,
                    "メモ": row[5] if row[5] else ""
                })
            except:
                continue
    
    return pd.DataFrame(data)

df = load_data()

if df.empty:
    st.error("データが見つかりません")
    st.stop()

# サイドバー：フィルタ
st.sidebar.header("⚙️ 表示設定")
mode = st.sidebar.radio("表示モード", ["日別", "期間指定", "全期間"], index=1)

if mode == "日別":
    target_date = st.sidebar.date_input("日付", value=df["日付"].max())
    filtered_df = df[df["日付"] == target_date]
    title_suffix = f"({target_date})"
elif mode == "期間指定":
    col1, col2 = st.sidebar.columns(2)
    start_date = col1.date_input("開始", value=df["日付"].min())
    end_date = col2.date_input("終了", value=df["日付"].max())
    filtered_df = df[(df["日付"] >= start_date) & (df["日付"] <= end_date)]
    title_suffix = f"({start_date} 〜 {end_date})"
else:
    filtered_df = df
    title_suffix = "(全期間)"

if filtered_df.empty:
    st.warning("指定期間にデータがありません")
    st.stop()

# タスク別集計
task_time = defaultdict(float)
for _, row in filtered_df.iterrows():
    task_time[row["タスク"]] += row["分"]

# カラーパレット
distinct_colors = [
    '#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8',
    '#F7DC6F', '#BB8FCE', '#85C1E2', '#F8B739', '#52B788',
    '#E07A5F', '#81B29A', '#F2CC8F', '#A8DADC', '#E63946'
]

# サマリー表示
st.subheader(f"📈 工数サマリー {title_suffix}")
col1, col2, col3 = st.columns(3)
total_hours = sum(task_time.values()) / 60
col1.metric("総工数", f"{total_hours:.1f} 時間")
col2.metric("タスク数", len(task_time))
col3.metric("記録日数", filtered_df["日付"].nunique())

st.markdown("---")

# グラフ表示
col_left, col_right = st.columns(2)

with col_left:
    st.subheader(f"タスク別工数 {title_suffix}")
    fig1, ax1 = plt.subplots(figsize=(6, 5), facecolor='#1a1a2e')
    ax1.set_facecolor('#16213e')
    
    tasks = list(task_time.keys())
    times = [task_time[t]/60 for t in tasks]
    colors = [distinct_colors[i % len(distinct_colors)] for i in range(len(tasks))]
    
    ax1.barh(tasks, times, color=colors)
    ax1.set_xlabel('時間 (h)', color='#f1f5f9', fontsize=11)
    ax1.tick_params(colors='#f1f5f9', labelsize=10)
    ax1.spines['bottom'].set_color('#94a3b8')
    ax1.spines['left'].set_color('#94a3b8')
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax1.grid(axis='x', color='#2d3748', linestyle='--', linewidth=0.5, alpha=0.7)
    
    st.pyplot(fig1)

with col_right:
    st.subheader(f"タスク割合 {title_suffix}")
    fig2, ax2 = plt.subplots(figsize=(6, 5), facecolor='#1a1a2e')
    ax2.set_facecolor('#16213e')
    
    sizes = [task_time[t]/60 for t in tasks]
    
    def autopct_format(pct):
        return f'{pct:.1f}%' if pct > 3 else ''
    
    wedges, texts, autotexts = ax2.pie(
        sizes,
        labels=tasks,
        autopct=autopct_format,
        colors=colors,
        textprops={'color': '#ffffff', 'fontsize': 10, 'weight': 'bold'},
        startangle=90,
        pctdistance=0.85
    )
    
    for autotext in autotexts:
        autotext.set_color('#000000')
        autotext.set_fontsize(11)
        autotext.set_weight('bold')
    
    for text in texts:
        text.set_fontsize(9)
    
    st.pyplot(fig2)

# データテーブル表示
st.markdown("---")
st.subheader("📋 詳細データ")
display_df = filtered_df.copy()
display_df["時間"] = (display_df["分"] / 60).round(1)
display_df = display_df[["日付", "開始", "終了", "タスク", "時間", "メモ"]]
st.dataframe(display_df, use_container_width=True, height=400)

# フッター
st.markdown("---")
st.caption("🔄 データは60秒ごとに自動更新されます")